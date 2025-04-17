import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
import os
import numpy as np
import pandas as pd
import random
import seaborn as sns
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
from scipy.ndimage import gaussian_filter1d


# Function to open a pop-up window for file selection (multiple CSV files)
def load_csv_files():
    root = tk.Tk()
    root.withdraw()  # Hides the main window
    response = messagebox.askyesno("Load Data", "Would you like to load existing data?")

    if response:
        # Load existing file
        filepath = filedialog.askopenfilename(title="Select an existing file", filetypes=[("Excel files", "*.xlsx")])
        if filepath:
            df = pd.read_excel(filepath)  # Load Excel file as DataFrame
            paths = df.values.tolist()  # Each line of the DataFrame becomes a path (list of values)

            if isinstance(paths, pd.DataFrame) or isinstance(paths[0], pd.Series):
                paths = [list(path) for path in df.values]  # Convert to lists
            paths = remove_nan_from_paths(paths)

            print(f"Loaded existing data from {filepath}")

            return paths, True, [filepath]  # Return paths as lists
    else:
        # Select new CSV files
        csv_files = filedialog.askopenfilenames(title="Select new CSV files", filetypes=[("CSV files", "*.csv")])
        data = [pd.read_csv(f) for f in csv_files]
        print(f"Loaded new CSV files: {csv_files}")
        return data, False, csv_files

def remove_nan_from_paths(paths):
    cleaned_paths = []
    for path in paths:
        # Remove NaN values from the path
        cleaned_path = [val for val in path if not np.isnan(val)]
        cleaned_paths.append(cleaned_path)
    return cleaned_paths

# Function for detecting local maxima
def detect_local_maxima(data):
    local_maxima = np.zeros_like(data, dtype=bool)

    # Run through each element (without border areas)
    for i in range(1, data.shape[0] - 1):
        for j in range(1, data.shape[1] - 1):
            # Extract the 3x3 window around the current element
            window = data[i - 1:i + 2, j - 1:j + 2]
            center_value = data[i, j]

            # Check whether the center point is larger than all its neighbors
            if center_value == np.max(window):
                local_maxima[i, j] = True

    return local_maxima


# Pop-up for sorting options
def sorting_options_popup():
    root = tk.Tk()
    root.withdraw()  # Hide window

    # Frage nach Sortierung (Ja/Nein)
    sort_response = messagebox.askyesno("Sorting Options", "Would you like to sort the data?")
    sort_by = None
    position = None

    if sort_response:
        # Question whether to sort by maximum or minimum
        sort_by = simpledialog.askstring("Sort By", "Sort by Global Maximum or Global Minimum?", initialvalue="Minimum")

        # Question about the position of the maximum/minimum in the histogram
        position = simpledialog.askstring("Position",
                                          "Where should the global maximum/minimum be positioned? (left, center, right)",
                                          initialvalue="center")

    return sort_response, sort_by, position

# Function for saving the paths
def save_paths_popup(data):
    root = tk.Tk()
    root.withdraw()  # Fenster ausblenden

    # Select storage location for the paths
    filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], title="Save Paths")



    if filepath:
        # Create a DataFrame from the paths
        df = pd.DataFrame(data)  # Save paths as lines (each line is a path)
        df.to_excel(filepath, index=False)
        print(f"Paths saved at {filepath}")
    else:
        print("Save cancelled")

# Function for drawing the heat map
def plot_heatmap(data, maxima_coords):
    fig, ax = plt.subplots()
    heatmap = ax.imshow(data, cmap='viridis', aspect='auto')

    # Mark the maxima
    for coord in maxima_coords:
        ax.add_patch(Rectangle((coord[1] - 0.5, coord[0] - 0.5), 1, 1, fill=False, edgecolor='red', lw=2))

    # Interactive selection of fields
    selected_cells = []

    def onclick(event):
        if event.inaxes == ax:
            x, y = int(event.xdata), int(event.ydata)
            if (y, x) in selected_cells:
                # Undo if the field is already selected
                selected_cells.remove((y, x))
                ax.add_patch(Rectangle((x-0.5, y-0.5), 1, 1, fill=False, edgecolor='white', lw=2))  # Removes the marker (sets it to white)
            else:
                selected_cells.append((y, x))
                ax.add_patch(Rectangle((x-0.5, y-0.5), 1, 1, fill=False, edgecolor='blue', lw=2))  # Set marker
            plt.draw()

    cid = fig.canvas.mpl_connect('button_press_event', onclick)
    plt.show()

    return selected_cells

# Function for checking neighboring fields and saving
def check_and_save_paths(selected_cells, data, excel_writer, sheet_name, maxima_coords):
    while True:
        try:
            for i in range(1, len(selected_cells)):
                y0, x0 = selected_cells[i-1]
                y1, x1 = selected_cells[i]
                if not ((abs(x1 - x0) <= 1 and abs(y1 - y0) <= 1)):
                    raise ValueError("Path is not contiguous! Fields must be adjacent (also diagonally).")
            break  # Valid path -> Exit loop
        except ValueError as e:
            print(f"Fehler: {e}, bitte versuche es erneut.")
            selected_cells = plot_heatmap(data, maxima_coords)

    # Entering the numerical values of the selected fields
    values = [data[y, x] for y, x in selected_cells]

    # Saving the values in the Excel file on the first spreadsheet
    df = pd.DataFrame([values])
    df.to_excel(excel_writer, sheet_name=sheet_name, index=False, startrow=excel_writer.sheets[sheet_name].max_row)


def position_paths(paths, position):
    if position == 'left':
        print("Positioning global min/max on the left...")
        return paths
    elif position == 'center':
        print("Positioning global min/max in the center...")
        return [np.roll(path, len(path) // 2) for path in paths]
    elif position == 'right':
        print("Positioning global min/max on the right...")
        return [np.roll(path, -1) for path in paths]

# Sorting function based on the user selection
def sort_paths(paths, sort_by):
    if sort_by.lower() == 'maximum':
        return [np.roll(path, -np.argmax(path)) for path in paths]
    else:
        return [np.roll(path, -np.argmin(path)) for path in paths]

def find_minimum_range(path,sort_by, threshold=10):

    min_index = sort_paths(path, sort_by)  # Index of the global minimum
    min_value = path[min_index]

    # Range to the left of the minimum (incl. cyclical logic)
    left = min_index
    while left > 0 and path[left - 1] <= min_value + threshold:
        left -= 1
    if left == 0 and path[-1] <= min_value + threshold:  # Cyclical extension left
        left = len(path) - 1

    # Range to the right of the minimum (incl. cyclical logic)
    right = min_index
    while right < len(path) - 1 and path[right + 1] <= min_value + threshold:
        right += 1
    if right == len(path) - 1 and path[0] <= min_value + threshold:  # Cyclic extension right
        right = 0

    return left, right

def normalize_paths(paths):
    normalized_paths = []
    for path in paths:
        path_length = len(path)
        # Normalize x-values from 0 to 1
        x_vals = np.linspace(0, 1, path_length)
        # Normalize the y-values (signal values)
        y_vals = (np.array(path) - np.min(path)) / (np.max(path) - np.min(path))
        # Save (x, y) pairs
        normalized_path = [(x, y) for x, y in zip(x_vals, y_vals)]
        normalized_paths.append(normalized_path)
    return normalized_paths

def normalize_path(path):
    min_val = np.min(path)
    max_val = np.max(path)
    return (path - min_val) / (max_val - min_val)


def stretch_paths_to_fit(paths, target_length=100):
    stretched_paths = []

    for path in paths:
        path_length = len(path)

        # Generate the x-values (0 to 1) for the current path
        x_vals = np.linspace(0, 1, path_length)

        # Interpolation of the y-values (signal values) to the new target length (target_length)
        stretched_x_vals = np.linspace(0, 1, target_length)
        stretched_y_vals = np.interp(stretched_x_vals, x_vals, path)  # Interpolation

        stretched_paths.append(stretched_y_vals)

    return stretched_paths

# Function to smooth a single line
def smooth_row(row, sigma=1):
    return gaussian_filter1d(row, sigma=sigma)

def mark_maxima_by_value(ax, maxima_values, path_data, row_index, num_rows):
    for j, value in enumerate(path_data):
        if value in maxima_values:
            # Mark maxima in the heat map with red vertical lines
            ax.axvline(x=j, color='red', linestyle='--', lw=1, ymin=row_index / num_rows, ymax=(row_index + 1) / num_rows)


def list_to_string(lst):
    return ', '.join(map(str, lst))


def detect_local_maxima_path(path):
    maxima_indices = []
    for i in range(1, len(path) - 1):
        if path[i - 1] < path[i] > path[i + 1]:
            maxima_indices.append(i)
    return maxima_indices

# Function for heatmap display of paths and maxima
def plot_heatmap_with_histogram(postioned_paths, maxima_coords, max_len, data):

    #Create the histogram and heatmap in reverse order
    fig, ax = plt.subplots(2, 1, figsize=(10, 8), gridspec_kw={'height_ratios': [1, 2]})

     # Route the paths and plot the path lines
    stretched_paths = []
    for path in postioned_paths:
        if len(path) != max_len:
            print(f"Warning: Path length {len(path)} does not match max_len {max_len}. Stretching path...")
            # Stretching the path if it does not have the length of max_len
            x_vals = np.linspace(0, 1, len(path))
            stretched_path = np.interp(np.linspace(0, 1, max_len), x_vals, path)
        else:
            stretched_path = path

        stretched_paths.append(stretched_path)
        ax[0].plot(np.linspace(0, 1, max_len), stretched_path, color='gray', alpha=0.5)

    # Calculate and plot average path
    avg_path = np.mean(stretched_paths, axis=0)
    ax[0].plot(np.linspace(0, 1, max_len), avg_path, color='black', lw=2)  # Average line in black
    ax[0].set_title("Signal Intensity Histogram (Paths)")

    # Create heatmap data
    heatmap_data = np.array(postioned_paths)
    sns.heatmap(np.array([smooth_row(row) for row in heatmap_data]), ax=ax[1], cmap='viridis', cbar=False)

    # Let the Y-axis start at 1
    ax[1].set_yticks(np.arange(1, len(postioned_paths) + 1))
    ax[1].set_yticklabels(np.arange(1, len(postioned_paths) + 1))

    # Find and draw maxima for each path
    for i, filled_paths in enumerate(postioned_paths):
        maxima_indices = detect_local_maxima_path(filled_paths)  # Local maxima for the current path
        for maxima_idx in maxima_indices:
            # Mark the maxima on the heatmap (cell is highlighted)
            ax[1].add_line(plt.Line2D([maxima_idx + 0.5, maxima_idx + 0.5], [i, i + 1], color='red', linestyle='--', lw=2))


    ax[1].set_title("Heatmap of Signal Intensity (Paths)")
    plt.tight_layout()
    plt.show()


# Function for plotting the legend separately
def plot_separate_legend():
    fig, ax = plt.subplots(figsize=(6, 1))
    fig.subplots_adjust(bottom=0.5)

    # Dummy heatmap for creating the legend
    cmap = plt.get_cmap('viridis')
    norm = plt.Normalize(vmin=0, vmax=1)

    fig.colorbar(plt.cm.ScalarMappable(norm=norm, cmap=cmap), cax=ax, orientation='horizontal')
    plt.show()



# Main function for processing CSV files
def process_csv_files():
    data, load_existing_data, file_paths = load_csv_files()

    # If existing data has been loaded, skip the path creation
    if load_existing_data:
        folder_path = os.path.dirname(file_paths[0])
        all_path_data=data
        print("Skipping path generation and filling steps, as existing data was loaded.")
    else:

        folder_path = os.path.dirname(file_paths[0])
        excel_path = os.path.join(folder_path, 'selected_paths.xlsx')

        # Excel file for saving the paths
        excel_writer = pd.ExcelWriter(excel_path, engine='openpyxl')

        # Create the first spreadsheet for the path display
        wb = excel_writer.book
        path_sheet = wb.create_sheet(title="Paths")
        excel_writer.sheets['Paths'] = path_sheet

        max_count_per_file = []  # List for saving the maximum number per file
        all_path_data = []  # List for saving all paths

        for file_path in file_paths:
            # Load CSV file
            if file_path.endswith('.csv'):
                df = pd.read_csv(file_path, index_col=0)
            else:
                df = pd.read_excel(file_path, index_col=0)

            # Extract the numerical data from the DataFrame (without labeling)
            data = df.iloc[:, 1:].values

            # Determining the local maxima
            local_maxima = detect_local_maxima(data)

            # Extract coordinates of the maxima
            maxima_coords = np.argwhere(local_maxima)

            # Display heatmap and select fields
            selected_cells = plot_heatmap(data, maxima_coords)

            # Create a new spreadsheet for each file
            sheet_name = file_path.split("/")[-1].split(".")[0]  # DFile name as sheet name
            wb.create_sheet(title=sheet_name)
            excel_writer.sheets[sheet_name] = wb[sheet_name]

            # Transfer the data to the corresponding spreadsheet
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
                for c_idx, value in enumerate(row, 1):
                    wb[sheet_name].cell(row=r_idx + 1, column=c_idx, value=value)

            # Mark the maxima in red
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            for coord in maxima_coords:
                wb[sheet_name].cell(row=coord[0] + 2, column=coord[1] + 2).fill = red_fill

            # Mark the selected fields in blue
            blue_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
            for y, x in selected_cells:
                wb[sheet_name].cell(row=y + 2, column=x + 2).fill = blue_fill

            # Check and save path
            try:
                check_and_save_paths(selected_cells, data, excel_writer, 'Paths', maxima_coords)
            except ValueError as e:
                print(f"Fehler: {e}")

            # Collect path values for later heatmap display
            path_values = [data[y, x] for y, x in selected_cells]
            all_path_data.append(path_values)


        save_paths_popup(all_path_data)

    max_len = max(len(path) for path in all_path_data)


    sort_response, sort_by, position = sorting_options_popup()

    if sort_response:
        sorted_paths = sort_paths(all_path_data, sort_by)
        positioned_paths = position_paths(sorted_paths, position)
    else:
        positioned_paths= all_path_data


    normalized_paths = [normalize_path(positioned_paths) for positioned_paths in positioned_paths]

    stretched_paths = stretch_paths_to_fit(normalized_paths)

    local_maxima_indices = [detect_local_maxima_path(path) for path in stretched_paths]


    # Create the heatmap and the histogram for the paths
    plot_heatmap_with_histogram(np.array(stretched_paths), local_maxima_indices, max_len, data)

    # Write the number of maxima per file in the last column
    max_count_df = pd.DataFrame(max_count_per_file, columns=["Maxima_Count"])
    max_count_df.to_excel(excel_writer, sheet_name="Paths", startcol=len(df.columns) + 2, index=False)

    # Save Excel file
    excel_writer.close()
    print("The selected paths and data have been saved successfully.")

    # Separates Fenster für die Legende anzeigen
    plot_separate_legend()


# Prozess starten
process_csv_files()