import cv2
import numpy as np
import os
import json
import matplotlib.pyplot as plt
import matplotlib.lines as mlines
import geopandas as gpd
import pygeoops
import pandas as pd
import ast
from shapely.geometry import LineString, Point
from matplotlib.widgets import Slider, Button
from shapely.ops import nearest_points
from shapely.affinity import rotate
from shapely.affinity import scale
from shapely.affinity import translate
from scipy.spatial import distance
from scipy.ndimage import uniform_filter1d
from shapely.ops import split, linemerge
from scipy.interpolate import splprep, splev
from scipy.interpolate import interp1d
from scipy.ndimage import gaussian_filter1d
from skimage.morphology import closing, remove_small_objects, square
from skimage import img_as_bool
from skimage import io, img_as_ubyte
from skimage.feature import peak_local_max
from skimage.filters import gaussian
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from tkinter import filedialog
from tkinter import Tk
from skimage.morphology import medial_axis
from skimage import io
from collections import Counter
from tkinter.filedialog import askdirectory

# --- Funktionen ---
pixel_size=0.100

class ContourWithAttributes:
    def __init__(self, contour, contour_id, centerline=None, image_name = None):
        self.contour_id = contour_id
        self.contour = contour
        self.image_name = image_name
        self.length = cv2.arcLength(contour, True) * pixel_size
        self.area = cv2.contourArea(contour) * (pixel_size ** 2)
        self.width = self.area / self.length if self.length != 0 else 0
        self.centerline= centerline
        self.scale_factor = max(self.length, self.width)
        self.signals = []


def select_input_folder(title):
    root = Tk()
    root.withdraw()  # Hides the main window
    folder_path = askdirectory(title=title)
    if not folder_path:
        raise ValueError("No folder selected")
    return folder_path



    # Dialog für die Auswahl der CSV-Dateien
    print("Please select the corresponding CSV files.")
    csv_files = filedialog.askopenfilenames(
        title="Select CSV Files",
        filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
    )
    if not csv_files:
        raise ValueError("No CSV files selected.")
    #print(f"Ausgewählte CSV-Dateien: {csv_files}")

    # Überprüfen der CSV-Dateinamen
    track_stats_files = [f for f in csv_files if "Track statistics_" in os.path.basename(f)]
    links_in_tracks_files = [f for f in csv_files if "Links in tracks statistics_" in os.path.basename(f)]

    if not track_stats_files:
        raise ValueError("No file with 'Track statistics_' in the name found.")
    if not links_in_tracks_files:
        raise ValueError("No file with 'Links in tracks statistics_' in its name found.")

    #print(f"Gefundene 'Track statistics_' Dateien: {track_stats_files}")
    #print(f"Gefundene 'Links in tracks statistics_' Dateien: {links_in_tracks_files}")

    # Dialog für den Ausgabeordner
    print("Please select the output folder.")

    output_folder = filedialog.askdirectory(title="Select Output Folder")
    if not output_folder:
        raise ValueError("No output folder selected.")
    #print(f"Ausgabeordner: {output_folder}")

    return {
        "track_stats_files": track_stats_files,
        "links_in_tracks_files": links_in_tracks_files,
        "output_folder": output_folder
    }

def read_csv_files(input_folder):
    track_stats_path = [f for f in os.listdir(input_folder) if "Track statistic" in f][0]
    links_stats_path = [f for f in os.listdir(input_folder) if "Links in tracks statistic" in f][0]

    track_stats = pd.read_csv(os.path.join(input_folder, track_stats_path), sep=None, engine='python')
    links_stats = pd.read_csv(os.path.join(input_folder, links_stats_path), sep=None, engine='python')

    return track_stats, links_stats

def categorize_tracks(track_displacement, categories):
    """
    Categorizes the value of track_displacement based on the defined categories.

    Parameters:
        track_displacement (float): The value to be categorized.
        categories (dict): A dictionary that contains the category names as keys and the limit values (min, max) as values.

    Returns:

        str: The name of the category to which the value belongs, or 'Unknown' if no category matches.
    """
    for category, (min_val, max_val) in categories.items():
        if min_val <= track_displacement <= max_val:
            return category  # Returns the name of the category
    return 'Unknown'

def plot_slider_and_get_thresholds(min_val, max_val):
    # Create only one empty figure for the slider
    fig = plt.figure(figsize=(6, 3))
    plt.subplots_adjust(left=0.1, bottom=0.25, top=0.8)

    # Slider for 'Static/Mobile'
    ax_static = plt.axes([0.1, 0.4, 0.8, 0.03], facecolor='lightgoldenrodyellow')
    s_static = Slider(ax_static, 'Static/Mobile', min_val, max_val, valinit=min_val)

    # Slider for 'Mobile/Fast Mobile'
    ax_mobile = plt.axes([0.1, 0.3, 0.8, 0.03], facecolor='lightgoldenrodyellow')
    s_mobile = Slider(ax_mobile, 'Mobile/Fast Mobile', min_val, max_val, valinit=max_val)

    # Button to confirm selection
    ax_button = plt.axes([0.4, 0.1, 0.2, 0.08])
    button_ok = Button(ax_button, 'OK')

    # Close the image by clicking on the OK button
    def on_button_click(event):
        plt.close(fig)

    button_ok.on_clicked(on_button_click)

    # Show only the window with the sliders and the button
    plt.show()

    # Return of the current slider values
    return s_static.val, s_mobile.val


def process_tracks(track_stats_df, num_categories):


    if num_categories == 1:
        categories = {'All': (track_stats_df.iloc[:, 11].min(), track_stats_df.iloc[:, 11].max())}
    elif num_categories == 2:
        min_val = track_stats_df.iloc[:, 11].min()
        max_val = track_stats_df.iloc[:, 11].max()
        static_threshold = plot_slider_and_get_thresholds(min_val, max_val)[0]
        categories = {'Static': (0, static_threshold), 'Mobile': (static_threshold, float('inf'))}
    elif num_categories == 3:
        min_val = track_stats_df.iloc[:, 11].min()
        max_val = track_stats_df.iloc[:, 11].max()
        static_threshold, mobile_threshold = plot_slider_and_get_thresholds(min_val, max_val)
        categories = {
            'Static': (0, static_threshold),
            'Mobile': (static_threshold, mobile_threshold),
            'Fast Mobile': (mobile_threshold, float('inf'))
        }
    else:
        raise ValueError("Number of categories must be 1, 2, or 3.")

    return categories

def extract_and_number_contours(input_folder, output_folder):

    fluorescence_files = [f for f in os.listdir(input_folder) if f.endswith('.png')]

    contour_folder = os.path.join(output_folder, 'contours')
    os.makedirs(contour_folder, exist_ok=True)
    json_folder = os.path.join(output_folder, 'json')
    os.makedirs(json_folder, exist_ok=True)

    contours_with_attributes = []
    contour_id = 1

    for fluorescence_file in fluorescence_files:

        image_path = os.path.join(input_folder, fluorescence_file)

        filename = os.path.basename(image_path)
        original_image = cv2.imread(image_path)

        image_with_contours = np.ones_like(original_image) * 255

        image_name = os.path.basename(image_path)  # Extrahiere den Bildnamen

        # Load the binary image and invert it
        image = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
        inverted_image = cv2.bitwise_not(image)

        # Apply thresholding to obtain black bacterial cells on a white background
        _, thresh = cv2.threshold(inverted_image, 128, 255, cv2.THRESH_BINARY)

        # Find the contour of the cell
        contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)

        # Create a user-defined legend element for contours
        contour_legend = mlines.Line2D([], [], color='black', label='Konturen')

        # Create a legend element for the center line
        centerline_legend = mlines.Line2D([], [], color='red', label='Mittellinie')

        plt.legend(handles=[contour_legend, centerline_legend])

        # Create objects for the contours with different attributes
        for contour in contours:


            contour = np.squeeze(contour)

            contour_line = LineString(contour)

            # Calculate the circumference of the contour and create evenly distributed distances
            num_interpolated_points=20
            length = contour_line.length
            distances = np.linspace(0, length,len(contour) + (len(contour) - 1) * num_interpolated_points)

            # Interpolate new points along the contour line
            interpolated_points = [contour_line.interpolate(distance) for distance in distances]
            interpolated_contour = np.array([(point.x, point.y) for point in interpolated_points], dtype=np.float32)

            smoothed_contour = smooth_and_interpolate_contour(interpolated_contour)

            contour_attr = ContourWithAttributes(smoothed_contour, contour_id, image_name=image_name)
            contours_with_attributes.append(contour_attr)

            # Draw the outline in green color
            cv2.drawContours(image_with_contours, [contour], -1, (0, 255, 0), 2)

            # Number the contours in the picture
            M = cv2.moments(contour)
            if M["m00"] != 0:
                cX = int(M["m10"] / M["m00"])
                cY = int(M["m01"] / M["m00"])
            else:
                cX, cY = 0, 0
            cv2.putText(image_with_contours, str(contour_id), (cX, cY),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 2)

            output_filename = f'contour_attributes_{filename}.json'
            output_path = os.path.join(json_folder, output_filename)
            contour_data = {
                "Contour_ID": contour_id,
                "Length": contour_attr.length,
                "Width": contour_attr.width,
                "Area": contour_attr.area,
                "Polygon": contour_attr.contour.tolist()
            }
            with open(output_path, 'w') as json_file:
                json.dump(contour_data, json_file, indent=4)
            contour_id += 1


        contour_image_filename = f'contour_{filename}'
        contour_image_path = os.path.join(contour_folder, contour_image_filename)
        cv2.imwrite(contour_image_path, image_with_contours)

    return contours_with_attributes, inverted_image

def smooth_and_interpolate_contour(contour, num_points=200):
    """
    Smoothes and interpolates a contour using splines.
        Args:   contour (ndarray): Original contour as Nx1x2 array.
                num_points (int): Number of points in the interpolated contour.

        Returns:   smoothed_contour (ndarray): The smoothed and interpolated contour as an Mx2 array.
    """
    contour_coords = contour.squeeze()
    x, y = contour_coords[:, 0], contour_coords[:, 1]

    # Spline preparation and smoothing
    tck, _ = splprep([x, y], s=1)  # Der Parameter 's' steuert den Glättungsgrad
    u_new = np.linspace(0, 1, num_points)
    x_new, y_new = splev(u_new, tck)

    # Combine the interpolated points into a new contour
    smoothed_contour = np.vstack((x_new, y_new)).T
    return smoothed_contour.astype(np.float32)

def find_centerline_from_extremes(contour, centerline_points=50, min_opposite_distance=5):
    """
     Calculates a centerline based on the farthest points and adjusts it to the middle course of the contour.
        Args:
            contour (ndarray): The contour as (x, y) coordinates.
            centerline_points (int): The number of points on the centerline.

        Returns:
            centerline (LineString): The calculated centerline as a shapely LineString.
            centerline_coords (list): The coordinates of the centerline.
    """
    # 1. Calculate the furthest points on the contour
    max_distance = 0
    start_point, end_point = None, None

    # Run through all possible pairs of dots and find the pair with the maximum distance
    for i, point1 in enumerate(contour):
        for point2 in contour[i + 1:]:
            dist = distance.euclidean(point1, point2)
            if dist > max_distance:
                max_distance = dist
                start_point, end_point = point1, point2

    # 2. Divide the contour into two halves based on the start and end points
    half1, half2 = [], []
    add_to_first_half = True

    for pt in contour:
        half1.append(pt) if add_to_first_half else half2.append(pt)
        if np.array_equal(pt, start_point) or np.array_equal(pt, end_point):
            add_to_first_half = not add_to_first_half


    # 3. Create a rough center line between the maximum distant points
    line_x = np.linspace(start_point[0], end_point[0], centerline_points)
    line_y = np.linspace(start_point[1], end_point[1], centerline_points)
    rough_centerline_coords = list(zip(line_x, line_y))

    # 4. Calculate the center point for each point on the rough center line
    adjusted_centerline_coords = []
    for pt in rough_centerline_coords:
        # Determine the opposite points from the two halves
        distances_half1 = distance.cdist([pt], half1)
        distances_half2 = distance.cdist([pt], half2)

        closest_half1 = half1[np.argmin(distances_half1[0])]
        closest_half2 = half2[np.argmin(distances_half2[0])]

        # Calculate the center of the two opposite points
        mid_x = (closest_half1[0] + closest_half2[0]) / 2
        mid_y = (closest_half1[1] + closest_half2[1]) / 2
        adjusted_centerline_coords.append((mid_x, mid_y))

    # Check whether there are enough points for the center line
    if len(adjusted_centerline_coords) < 2:
        raise ValueError("The calculated center line contains too few points.")

    # Create the final center line as a LineString
    centerline = LineString(adjusted_centerline_coords)
    return centerline, adjusted_centerline_coords

def project_signals_onto_centerline(centerline, contour, signal_positions):

    contour_polygon = LineString(contour)

    projected_signals = []
    for i, signal in enumerate(signal_positions) :

        if isinstance(signal, (tuple, list)) and len(signal) == 5:
            x, y, track_id, contour_id, image_id = signal
            orig_signal_point = Point(x, y)

        else:
            print(f"Ungültiges Signal an Position {i + 1}: {signal}")
            continue  # Überspringe ungültige Signale

        # Find the next point on the centerline to the signal
        nearest_on_centerline = nearest_points(centerline, orig_signal_point)[0]

        # Calculate the distance along the centerline from the starting point to the projection point
        distance_along_centerline = centerline.project(nearest_on_centerline)

        # Calculate the orthogonal distance between signal and projection
        orthogonal_distance = orig_signal_point.distance(nearest_on_centerline)

        # Add the projection information
        projected_signals.append((nearest_on_centerline.x, nearest_on_centerline.y, distance_along_centerline, orthogonal_distance, orig_signal_point, track_id, contour_id, image_id))


    return projected_signals

def straighten_centerline_and_relocate_signals(centerline, signal_data):
    """
    Transforms the curved center line into a straight line of the same length and adjusts the signal positions, to maintain their relative position to the centerline.

    Args:
         centerline (LineString): The original (curved) centerline.
         signal_data (list): List of signal data as (distance_along_centerline, orthogonal_distance, x, y) tuples.

    Returns:
        straight_centerline (LineString): The straight centerline with the same length as the curved line.
        relocated_signals (list): List of the newly calculated signal coordinates as (x, y)-tuples.
    """
    # Calculate the length of the curved center line
    total_length = centerline.length

    # Define the straight centerline with the same length along the x-axis
    start_point = centerline.coords[0]
    end_point = (start_point[0] + total_length, start_point[1])
    straight_centerline = LineString([start_point, end_point])

      # Calculate new positions for the signals based on the relative positions
    relocated_signals = []
    for i, (distance_along_centerline, orthogonal_distance,  proj_x, proj_y, orig_signal_point, track_id, contour_id, image_id) in enumerate(signal_data):

        # Find the point on the new, straight center line
        point_on_straight = straight_centerline.interpolate(distance_along_centerline)

        # Compare the y-coordinate of the signal with the y-coordinate of nearest_on_centerline
        if (orig_signal_point.y - proj_y)  < 0:
            orthogonal_distance = -orthogonal_distance

        # Verschiebe das Signal entlang des rechten Winkels zur geraden Mittellinie
        new_x = point_on_straight.x + orthogonal_distance * (end_point[1] - start_point[1]) / total_length
        new_y = point_on_straight.y + orthogonal_distance


        relocated_signals.append((new_x, new_y, orig_signal_point.x, orig_signal_point.y, track_id,contour_id, image_id))

    return straight_centerline, relocated_signals


def shift_to_origin(centerline, signals):
    """
    Moves the center line and the signal positions to the origin.
    """
    # Determine the center of the center line
    centerline_midpoint = centerline.interpolate(0.5, normalized=True)
    dx, dy = -centerline_midpoint.x, -centerline_midpoint.y

    # Move the center line
    shifted_centerline = translate(centerline, dx, dy)

    # Shift the signals relative to the center line and retain additional data
    shifted_signals = [
        (signal[0] + dx, signal[1] + dy, *signal[2:])  # All other values are retained here
        for signal in signals
    ]

    return shifted_centerline, shifted_signals

def normalize_centerline_length(centerline, signals):
    # Calculate the scaling factors for the normalization of the center line length
    target_length = 100
    scale_factor = target_length / centerline.length

    # Skaliere die Mittellinie
    normalized_centerline = scale(centerline, xfact=scale_factor, yfact=scale_factor, origin=centerline.centroid)


    # Check and make sure that 'signals' is a list
    if isinstance(signals, dict):
        signals = [signals]

    # Scale the signals and save both the scaled and the original coordinates
    normalized_signals = [
        {
            'Track_ID': signal['Track_ID'],
            'EDGE_X_LOCATION': signal['EDGE_X_LOCATION'] * scale_factor,
            'EDGE_Y_LOCATION': signal['EDGE_Y_LOCATION'],
            'TRACK_DISPLACEMENT': signal['TRACK_DISPLACEMENT'],
            'Category': signal['Category'],
            'Image': signal['Image'],
            'orig_Track_ID':signal['orig_Track_ID']
        }
        for signal in signals
    ]

    return normalized_centerline, normalized_signals

def detect_delimiter(file_path):
    """Detect the delimiter of a CSV file (comma or semicolon)."""
    with open(file_path, 'r') as file:
        first_line = file.readline()
        if ',' in first_line:
            return ','
        elif ';' in first_line:
            return ';'
        else:
            raise ValueError(f"Unknown delimiter in file: {file_path}")


def filter_signals_by_track(all_signals, all_tracks):
    """
     Filters signals from all_signals_df based on the track data in all_tracks_df.
         Args:
            all_signals_df (pd.DataFrame): DataFrame with the signal data. The second column contains the TRACK_ID.
            all_tracks_df (pd.DataFrame): DataFrame with the track data. The second column contains the TRACK_ID, the third column contains the value for the filter condition.

        Returns:
            pd.DataFrame: Filtered signal data from all_signals_df.
    """
    # Add column names if not available
    all_signals_df = pd.DataFrame(all_signals)
    signals_filtered = all_signals_df.iloc[:, [1, 6, 7, 12]]
    signals_filtered.columns = ['TRACK_ID', 'EDGE_X_LOCATION', 'EDGE_Y_LOCATION', 'Image']

    all_tracks_df = pd.DataFrame(all_tracks)
    tracks_filtered = all_tracks_df.iloc[:, [1, 2]]
    tracks_filtered.columns = ['TRACK_ID', 'Filter_Value']

    # Filter the TRACK_IDs based on the filter condition
    valid_tracks = tracks_filtered.loc[tracks_filtered['Filter_Value'] >= 3, 'TRACK_ID']
    #print(f"Gefilterte TRACK_IDs: {valid_tracks.tolist()}")  # Debugging

    # Filter the signals based on the valid TRACK_IDs
    filtered_signals = signals_filtered[signals_filtered['TRACK_ID'].isin(valid_tracks)]

    return filtered_signals

def save_tracks_to_excel(tracks, output_file, categories):
    """
     Saves the tracks in an Excel file, with each category having its own spreadsheet.
     Args:
        tracks (list): List of track data as dictionaries with the keys 'Track_ID', 'EDGE_X_LOCATION', 'EDGE_Y_LOCATION', 'TRACK_DISPLACEMENT'.
        output_file (str): The path to the output file.
        categories (dict): Categories with their limit values.
    """
    wb = Workbook()
    for category, (min_val, max_val) in categories.items():
        # Create a spreadsheet for the category
        ws = wb.create_sheet(title=category)
        ws.append(["Track_ID", "EDGE_X_LOCATION", "EDGE_Y_LOCATION", "TRACK_DISPLACEMENT","Original_Track_ID"])

        # Filter the tracks for this category and insert them
        for track in tracks:
            if min_val <= track["TRACK_DISPLACEMENT"] <= max_val:
                ws.append([
                    track["Track_ID"],
                    track["EDGE_X_LOCATION"],
                    track["EDGE_Y_LOCATION"],
                    track["TRACK_DISPLACEMENT"],
                    track["orig_Track_ID"]
                ])

    # Remove the default worksheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Save the Excel file
    wb.save(output_file)
    print(f"Tracks wurden in {output_file} gespeichert.")

def filter_signals_within_contour(all_signals_df, contour, contour_image_name, contour_id):
    # Extract the image number from the contour image name
    try:
        contour_image_id = int(contour_image_name.split('_')[-1].split('.')[0])
    except (IndexError, ValueError):
        raise ValueError(f"Ungültiger Bildname: {contour_image_name}")

    # Filter signals based on the image number
    signals_matching_image = all_signals_df[all_signals_df['Image'] == contour_image_id]

    # Initialize results
    signal_positions = []
    signals_in_contour = {}

    # Iterate through the filtered signals
    for _, signal in signals_matching_image.iterrows():
        x, y = float(signal['EDGE_X_LOCATION']) * 10, float(signal['EDGE_Y_LOCATION']) * 10
        track_id = int(signal['TRACK_ID'])
        image_id = int(signal['Image'])
        # Prüfe, ob das Signal innerhalb der Kontur liegt
        if cv2.pointPolygonTest(contour, (x, y), False) >= 0:
            position = (x, y, track_id, contour_id, image_id)
            signal_positions.append(position)
            signals_in_contour[(x, y)] = signal.to_dict()

    # Return of the filtered signals
    return signal_positions, signals_in_contour

def visualize_centerline_and_signals(centerline, contour, signal_positions, projected_signals, title):

    # Contour plot
    contour_x, contour_y = zip(*contour)
    plt.plot(contour_x, contour_y, color='blue', label='Kontur')

    # Centerline plot
    centerline_x, centerline_y = zip(*centerline.coords)
    plt.plot(centerline_x, centerline_y, color='green', label='Mittellinie')

    # Plotting original signal positions
    if signal_positions:
        orig_signal_x, orig_signal_y = zip(*[(x, y) for x, y, *_ in signal_positions])
        plt.scatter(orig_signal_x, orig_signal_y, color='orange', label='Originale Signalpositionen')

    # Plot projected signals (x and y coordinates only)
    if projected_signals:
        proj_signal_x, proj_signal_y = zip(*[(proj_x, proj_y) for proj_x, proj_y, *_ in projected_signals])
        plt.scatter(proj_signal_x, proj_signal_y, color='red', marker='x', label='Projektierte Signale')

    plt.title(title)
    plt.xlabel("X-Koordinate")
    plt.ylabel("Y-Koordinate")
    plt.legend()
    plt.axis("equal")
    plt.show()

def main():
    input_folder = select_input_folder(title="Select input-file")
    fluorescence_files = [f for f in os.listdir(input_folder) if f.endswith('.png')]
    csv_files = [f for f in os.listdir(input_folder) if f.endswith('.csv')]

    output_folder = select_input_folder(title="Select output-file")

    num_categories = int(input("Enter the number of categories (1, 2, or 3): "))

    contours_with_attributes, _ = extract_and_number_contours(input_folder, output_folder)
    # Separate the files based on the names
    all_signals_df = [f for f in csv_files if f.startswith("Links in tracks statistics")]
    all_tracks_df = [f for f in csv_files if f.startswith("Track statistics")]

    if all_signals_df is None:
        print("The 'Links in tracks statistics_' file could not be processed.")
        return

    # Initialize a list to collect all track data
    combined_tracks_df = pd.DataFrame()

    # Iterate over all track statistics files and merge the data
    for track_stats_file in all_tracks_df:
        track_stats_full_path = os.path.join(input_folder, track_stats_file)

        # Load the current track statistics file
        current_tracks_df = pd.read_csv(track_stats_full_path, sep=';')

        # Add the data to the combined DataFrame
        combined_tracks_df = pd.concat([combined_tracks_df, current_tracks_df], ignore_index=True)
    categories = process_tracks(combined_tracks_df, num_categories)

    all_adjusted_signals = []
    all_centerlines = []
    all_filtered_signals=[]

    # Iterate over each image and the corresponding CSV file
    for fluorescence_image_path, links_in_tracks_file, track_stats_file in zip(fluorescence_files, all_signals_df, all_tracks_df):

        links_in_tracks_full_path = os.path.join(input_folder, links_in_tracks_file)
        track_stats_full_path = os.path.join(input_folder, track_stats_file)

        # Load the CSV data for this image
        all_signals_df = pd.read_csv(links_in_tracks_full_path, sep=';')
        all_tracks_df = pd.read_csv(track_stats_full_path, sep=';')

        all_signals = all_signals_df.to_dict('records')  # Each line as a dictionary

        filtered_signals = filter_signals_by_track(all_signals, all_tracks_df)
        all_filtered_signals.append(filtered_signals)


        for contour_attr in contours_with_attributes:
            enhanced_signals = []

            centerline, centerline_coords = find_centerline_from_extremes(contour_attr.contour)
            contour_attr.centerline = centerline

            signal_positions, signals_in_contour = filter_signals_within_contour(filtered_signals, contour_attr.contour,
                                                                             contour_attr.image_name, contour_attr.contour_id)

            # Project the signals onto the center line of the contour
            projected_signals = project_signals_onto_centerline(centerline, contour_attr.contour, signal_positions)

            for proj_x, proj_y, distance_along_centerline, orthogonal_distance,orig_signal_point, track_id, contour_id, image_id in projected_signals:
                    enhanced_signals.append((distance_along_centerline, orthogonal_distance, proj_x, proj_y, orig_signal_point, track_id, contour_id, image_id))


        # Smooth centerline and adjust signals
            smoothed_centerline, relocated_signals = straighten_centerline_and_relocate_signals(centerline, enhanced_signals)

        #Move centerline and signals to the origin
            shifted_centerline, shifted_signals = shift_to_origin(smoothed_centerline, relocated_signals)

            all_centerlines.append(shifted_centerline)
            all_adjusted_signals.append(shifted_signals)


    tracks = []

        # Iteration over the entries in normalized_signals

    for signal_group in all_adjusted_signals:

            for signal in signal_group:


                track_id = signal[4]
                image_id = signal[6]

                if isinstance(track_id, (list, tuple, np.ndarray)):
                    track_id = track_id[4]

                if isinstance(image_id, (list, tuple, np.ndarray)):
                    image_id = image_id[6]


                # Filter the line in combined_tracks_df based on TRACK_ID and image
                filtered_row = combined_tracks_df[
                    (combined_tracks_df["TRACK_ID"] == track_id) &
                    (combined_tracks_df["Image"] == image_id)
                    ]

                # Check whether a value has been found before extracting it
                if not filtered_row.empty:
                    track_displacement = filtered_row["TRACK_DISPLACEMENT"].values[0]
                else:
                    track_displacement = None
                    print(f"No suitable TRACK_DISPLACEMENT value for TRACK_ID {track_id} and Image {image_id}")

                # Categorization based on TRACK_DISPLACEMENT
                category = categorize_tracks(track_displacement, categories)

                # Track-Daten speichern
                track_data = {
                   "Track_ID": track_id,
                    "EDGE_X_LOCATION": signal[0],  # X-Koordinate
                    "EDGE_Y_LOCATION": signal[1],  # Y-Koordinate
                    "TRACK_DISPLACEMENT": track_displacement,
                    "Category": category,
                    "Contour": signal[5],
                    "Image" : signal[6],
                    "orig_Track_ID": track_id
                }
                tracks.append(track_data)


    # Link the center lines with the track IDs
    centerline_map = {}
    for contour, centerline in zip(contours_with_attributes, all_centerlines):
        centerline_map[contour.contour_id] = centerline

    normalized_centerlines = []
    normalized_signals = []

    for track in tracks:
        track_contour_id = track['Contour']  # ID des Bildes, zu dem der Track gehört
        centerline = centerline_map.get(track_contour_id)

        if not centerline:
            continue

        # Normalize the center line and signals
        norm_centerline, norm_signals = normalize_centerline_length(centerline, track)

        normalized_centerlines.append(norm_centerline)
        normalized_signals.append(norm_signals)

    # Merging the signals from all data sets
    final_normalized_signals = [signal for group in normalized_signals for signal in group]

    unique_tracks = []
    current_track_id = 0

    # Auxiliary variables to track the last Track_ID value and the image
    last_track_id = None
    last_image = None

    for track in final_normalized_signals:

        if track['Track_ID'] != last_track_id or track['Image'] != last_image:
            current_track_id += 1

        # Update the track with the new ID
        updated_track = track.copy()
        updated_track['Track_ID'] = current_track_id
        unique_tracks.append(updated_track)

        # Update the auxiliary variables
        last_track_id = track['Track_ID']
        last_image = track['Image']

    # Storage location of the output file
    output_file = os.path.join(output_folder, 'tracks_categorized.xlsx')

    # Save the tracks in the Excel file
    save_tracks_to_excel(unique_tracks, output_file, categories)

if __name__ == "__main__":
    main()